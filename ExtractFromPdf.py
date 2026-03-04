#!/usr/bin/env python3
"""
Script d'extraction de texte depuis un PDF avec pdfplumber
"""
import os
import re
import shutil
import sys
from pathlib import Path
from config import config
import pdfplumber

from Excel_utils import get_column_index

from ExcelToDict import excel_to_dict
from openpyxl import load_workbook
from openpyxl import Workbook
from setup_logger import setup_logger

logger = setup_logger(__name__)
# Configuration basique du logger


from ExtractPictoFromPdf import analyser_fds
MAX_PAGES_TO_SCAN = 50
MIN_LINES_WITHOUT_INFORMATIONS = 10
#To call external API to recognize picto from pictures


dictMention = {}
mentionInFile= {}


def extraire_texte_pdf(pathPDF):
    global mentionInFile
    """
    Extrait tout le texte d'un PDF

    Args:
        pathPDF: Chemin vers le fichier PDF

    Returns:
        str: Texte extrait du PDF
    """
    texte_complet = []
    logger.debug("extraire_texte_pdf......")
    try:
        with pdfplumber.open(pathPDF) as pdf:
            logger.debug(f"📄 Nombre de pages: {len(pdf.pages)}")
            numParaph=''
            interest = False
            interestTransport = False
            interestPicto = 0
            interestDanger = 0
            interestComplement = 0
            interestPrudence = 0
            interestAvertissement = 0
            expectContient = False
            pictos = []
            dangers = []
            complements = []
            prudences = []
            contients = []
            mentionInFile = {}
            last_prudence = ""
            avertissement=''
            transport=''
            # Parcourir toutes les pages
            for num_page, page in enumerate(pdf.pages, start=1):
                logger.debug(f"⚙️  Extraction page {num_page}...")
                if (num_page > MAX_PAGES_TO_SCAN):
                    break
                # Extraire le texte de la page
                texte = page.extract_text()

                p_picto = re.compile('PICTO', re.IGNORECASE)
                if texte:
                    lignes = texte.split('\n')
                    for ligne in lignes:
                        ligne = ligne.strip()
                        logger.debug(f"================{ligne}=====================")
                        numParaph=''
                        if (m := re.match(r'[^0-9]{0,3}[0-9]+\.[0-9]+', ligne, re.IGNORECASE)):
                        # pour traiter "· 2.1 Classificat...."
                        # pour traiter " .2.1 Classificat...."
                            numParaph=m.group()
                            numParaph=re.sub(r'\.$','',numParaph)
                            numParaph = re.sub(r'[^0-9\.]', '', numParaph)
                            numParaph = re.sub(r'^[^0-9]', '', numParaph)
                        if (numParaph == '14.2'):
                            logger.debug(f"interestTransport False due to :{ligne}")
                            interestTransport = False
                        if (not interestTransport and numParaph == '14.1'):
                            logger.debug(f"interestTransport True due to :{ligne}")
                            interestTransport = True
                        if (interestTransport):
                            if ((re.search(r'ADR',prevLine) or re.search(r'ADR',ligne)) and
                                (re.search(r'ONU',prevLine) or re.search(r'ONU',ligne) or re.search(r' UN ',ligne) )):
                                if (m := re.search(r'\d{4}',ligne)):
                                    transport='UN'+m.group()
        #                        if (not interest and (m := re.search('([^0-9.\-]*)([0-9.\- ]+).*TIQUETAGE', ligne, re.IGNORECASE))):
                        if (interest and numParaph == '2.3'):
                            logger.debug(f"interest False due to :{ligne}")
                            interest = False
                        if (not interest and  numParaph == '2.2'):
                            logger.debug(f"interest True due to :{ligne}")
                            interest =True
                        if (interest):
                            logger.debug(f"interest True")
                            if (t := re.search('PICTO', ligne,re.IGNORECASE)):
                                logger.debug(f"Picto =>  {t}")
                                interestPicto = MIN_LINES_WITHOUT_INFORMATIONS
                            if (interestPicto):
                                pictos_tmp=(re.findall(r'GHS\d+', ligne))
                                if (pictos_tmp):
                                    interestPicto = MIN_LINES_WITHOUT_INFORMATIONS
                                    for element in pictos_tmp:
                                        if not element in pictos:
                                            pictos.append(element)
                                else:
                                    interestPicto-= 1
                            if (interestAvertissement):
                                logger.debug(f"{avertissement} {ligne}")
                                if avertissement != "":
                                    interestAvertissement=0
                                else:
                                    avertissement=ligne
                            if (x := re.search("Mention d'avertissement(.*)", ligne)):
                                interestAvertissement=1
                                avertissement=x.group(1)
                                if (y := re.search(r': *(.*)',avertissement)):
                                    avertissement = y.group(1)
                            elif (x := re.search("d'avertissement(.*)", ligne)):
                                logger.debug(f"Seen <d'avertissemnt> alone in {ligne}")
                                avertissement=x.group(1)
                                if (y := re.search(r': *([^ ]+)',avertissement)):
                                    avertissement = y.group(1)
                                    logger.debug(f"avertissemnt <{avertissement}> found in {ligne}")
                                else:
                                    logger.debug(f"search avertissemnt in prevligne {prevLine}")
                                    if (y := re.search(r'Mention([ -:]+)(.+)',prevLine)):
                                        avertissement = y.group(2)
                            if (x := re.search("danger", ligne)):
                                interestDanger=MIN_LINES_WITHOUT_INFORMATIONS
                            if (interestDanger):
                                dangers_tmp=(re.findall(r'EUH\d+[A]?|H\d+', ligne))
                                if (dangers_tmp):
                                    interestDanger = MIN_LINES_WITHOUT_INFORMATIONS
                                    toadds = clean_mention(dangers_tmp)
                                    for element in toadds:
                                        if (not re.search(r'^EUH',element)):
                                            if not element in dangers:
                                                dangers.append(element)
                                                logger.debug(f"danger : {element} ")

                                else:
                                    interestDanger-=1
                            if (expectContient):
                                pattern = re.compile(r'[^A-Za-z0-9, :\-_()\'éèà/]+')
                                contient = pattern.sub('', ligne)
                                contient = re.sub(r'^ +','',contient)
                                if contient:
                                    contients.append(contient)
                                    expectContient=False
                            if ((re.search("(Contient|Composants dangereux)", ligne, re.IGNORECASE))):
                                tmp = re.split(r' *: *',ligne)
                                contient=tmp[1]
                                if (contient):
                                    contients.append(tmp[1])
                                else:
                                    expectContient=True

                            if ((re.search("(indications|informations) (compl|suppl)", ligne, re.IGNORECASE)) or
                                (re.search("Phrases EUH", ligne, re.IGNORECASE)) or
                                (re.search("Mentions de danger spécifiques", ligne, re.IGNORECASE))) :
                                interestComplement=MIN_LINES_WITHOUT_INFORMATIONS
                                logger.debug(f"Complement {ligne}")
                            if (interestComplement):
                                complements_tmp=(re.findall(r'EUH\d+[A]?|H\d+', ligne))
                                if (complements_tmp):
                                    interestComplement = MIN_LINES_WITHOUT_INFORMATIONS
                                    toadds=clean_mention(complements_tmp)
                                    for element in toadds:
                                        if (not element in complements):
                                            complements.append(element)
                                else:
                                    interestComplement-=1

                            if ((re.search("conseil", ligne, re.IGNORECASE)) or
                                 (re.search("en garde", ligne, re.IGNORECASE))):
                                interestPrudence=MIN_LINES_WITHOUT_INFORMATIONS
                                interestDanger=0
                                logger.debug(f"Prudence {ligne}")
                            if (interestPrudence):
                                logger.debug(f"...Prudence {interestPrudence}.... {ligne}")

                                prudences_tmp=(re.findall(r'P\d+', ligne))
                                if (prudences_tmp):
                                    toadds=clean_mention(prudences_tmp)
                                    for toadd in toadds:
                                        if not toadd in prudences:
                                            prudences.append(toadd)
                                            logger.debug(f"toadd : {toadd}")
                                            last_prudence = toadd
                                    mentionInFile[last_prudence] = ligne
                                    interestPrudence=MIN_LINES_WITHOUT_INFORMATIONS
                                else:
                                    if last_prudence:
                                        mentionInFile[last_prudence]+= ligne
                                    interestPrudence-=1
                        prevLine=ligne
                    else:
                        continue
                    break
                else:
                    logger.warning(f"⚠️  Aucun texte trouvé sur la page {num_page}")
            logger.debug(f"Pictos {pictos}, avertissement \"{avertissement}\" dangers {dangers} complements {complements} prudence {prudences}")
            for key, value in mentionInFile.items():
                print(f"{key} : {value}")
            return {
                'pictos' : pictos,
                'contients' : contients,
                'avertissement' : avertissement,
                'dangers': dangers,
                'complements': complements,
                'prudences': prudences,
                'transport': transport,
            }

    except FileNotFoundError:
        logger.error(f"❌ Erreur: Le fichier '{pathPDF}' n'existe pas")
        return None
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'extraction: {e}")
        return None
def clean_mention(mentions):
    results=[]
    nbMentions=len(mentions)
    while (nbMentions>0):
        restMention='+'.join(mentions)
        if (restMention in dictMention):
            results.insert(0,restMention)
            return results
        lastMention=mentions.pop()
        if (lastMention in dictMention):
            results.insert(0,lastMention)
    return(results)

def extraire_tableaux_pdf(pathPDF):
    """
    Extrait les tableaux d'un PDF

    Args:
        pathPDF: Chemin vers le fichier PDF

    Returns:
        list: Liste des tableaux extraits
    """
    tous_tableaux = []

    try:
        with pdfplumber.open(pathPDF) as pdf:
            for num_page, page in enumerate(pdf.pages, start=1):
                tableaux = page.extract_tables()

                if tableaux:
                    logger.debug(f"📊 {len(tableaux)} tableau(x) trouvé(s) sur la page {num_page}")
                    tous_tableaux.extend(tableaux)

        return tous_tableaux

    except Exception as e:
        logger.error(f"❌ Erreur lors de l'extraction des tableaux: {e}")
        return []


def update_sticker_file(fds , sheetName, fileSticker,key='yyy'):
    try:
        fileStickerTmp = re.sub(r'\.xlsm$',"_"+key+".xlsm",fileSticker)
        shutil.copy(fileSticker, fileStickerTmp)
        wb = load_workbook(fileStickerTmp)
    except Exception as e:
        print(f"Erreur : {e}", file=sys.stderr)
        sys.exit(1)
    ws = wb.worksheets[0]

    indexFds = get_column_index(ws, "FDS",exit_now=True)
    indexDanger = get_column_index(ws, "Mentions de danger",exit_now=True)
    indexAvertissement = get_column_index(ws, "Mention d'avertissement",exit_now=True)
    indexPrudence = get_column_index(ws, "Conseils de prudence",exit_now=True)
    indexContient = get_column_index(ws, "Contient",exit_now=True)
    indexModele = get_column_index(ws, "Modèle d'étiquette",exit_now=True)

    for i in range(2, ws.max_row + 1):
        rowfdsname = ws.cell(row=i, column=indexFds + 1).value
        logger.debug(f"Try to recognize{sheetName} in {rowfdsname}")
        if (ws.cell(row=i, column=indexFds + 1).value == sheetName):
            nb_pictos_sticker = 3
            Model = ws.cell(row=i, column=indexModele + 1).value
            if (match := re.search(r'(\d+)',Model)):
                nb_pictos_sticker = int(match[1])
            if ('pictos' in fds):
                j=1
                for picto in fds['pictos']:
                    indexPicto = get_column_index(ws,f"PICTO {j}",exit_now=True)
                    ws.cell(i, indexPicto + 1 ).value = picto
                    j = j + 1
                for k in range(j, nb_pictos_sticker + 1):
                    indexPicto = get_column_index(ws, f"PICTO {j}", exit_now=True)
                    ws.cell(i, indexPicto + 1).value = config['SETTINGS']['black_picto']
                    j = j + 1
            avertissement= ''
            if 'avertissement' in fds:
                avertissement =  fds['avertissement']
            ws.cell(row=i, column=indexAvertissement + 1).value = avertissement

            mentionDanger = ''
            if ('dangers' in fds):
                sep = ''
                for danger in fds['dangers']:
                    mentionDanger = sep + mentionDanger + danger + " " + dictMention[danger]
                    sep = ' '
            ws.cell(row=i, column=indexDanger + 1).value = mentionDanger

            mentionPrudence = ''
            if ('prudences' in fds):
                sep = ''
                for prudence in fds['prudences']:
                    mentionToApply = fromFileifVar(prudence)
                    mentionPrudence = sep + mentionPrudence + prudence + " " + mentionToApply
                    sep = ' '
            ws.cell(row=i, column=indexPrudence + 1).value = mentionPrudence

            mentionContient = ''
            if ('contients' in fds):
                sep = ''
                for contient in fds['contients']:
                    mentionContient = sep + mentionContient + contient
                    sep = ' '

            ws.cell(row=i, column=indexContient + 1).value = mentionContient

    wb.save(fileSticker)
def write_fds(fds , sheetName):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        wb = load_workbook(os.path.join(script_dir,config['PATHS']['pathFdsExcel']))
    except Exception as e:
        wb=Workbook()
    if (sheetName in wb.sheetnames):
        del wb[sheetName]
    if 'Sheet' in wb.sheetnames:
        ws = wb['Sheet']
        ws.title= sheetName
    else:
        ws=wb.create_sheet(sheetName)
    ws['A1'] = sheetName
    ws['A2'] = 'Pictogrammes'
    currentLine=3
    if ('pictos' in fds):
        for picto in fds['pictos']:
            ws['A'+str(currentLine)] = picto
            ws['B'+str(currentLine)] = dictMention[picto]
            currentLine+=1
    currentLine = 8
    ws['A' + str(currentLine)] = 'Contient'
    if ('contients' in fds):
        for contient in fds['contients']:
            ws['B'+str(currentLine)] = contient
            currentLine+=1
    currentLine = 9
    if ('transport' in fds):
        ws['A'+str(currentLine)] = "Numéro ONU ADR"
        ws['B'+str(currentLine)] = fds['transport']

    currentLine = 10
    if 'avertissement' in fds:
        ws['A'+str(currentLine)] = "Mention d'avertissement"
        ws['B'+str(currentLine)] = fds['avertissement']
    currentLine = 11
    ws['A' + str(currentLine)] = 'Mentions de danger'
    currentLine+=1
    if ('dangers' in fds):
        for danger in fds['dangers']:
            ws['A' + str(currentLine)] = danger
            ws['B' + str(currentLine)] = dictMention[danger]
            currentLine += 1
    currentLine = 30
    ws['A' + str(currentLine)] = 'Conseils de prudence'
    currentLine+=1
    if ('prudences' in fds):
        for prudence in fds['prudences']:
            ws['A' + str(currentLine)] = prudence
            ws['B' + str(currentLine)] = dictMention[prudence]
            currentLine += 1

    currentLine = 60
    ws['A' + str(currentLine)] = 'Mentions complémentaires'
    currentLine+=1
    if ('complements' in fds):
        logger.debug('complements found')
        for complement in fds['complements']:
            ws['A' + str(currentLine)] = complement
            if (complement in dictMention):
                ws['B' + str(currentLine)] = dictMention[complement]
            currentLine += 1

    wb.save(config['PATHS']['pathFdsExcel'])
def fromFileifVar(code):
    global mentionInFile
    result = dictMention[code]
    if not '…' in result:
        return result
    if (not code in mentionInFile):
        logger.error(f"{code} contains '…'not and not in mentionInFile")
        return result
    match = re.search(rf'{code}\s+(.*)', mentionInFile[code])
    if match:
        logger.debug(f"{code} in file seems to value {match.group(1)}")
        result = match.group(1)  # "
        matchEnd = re.search(r'^([^·]+)·(.*)',result)
        if matchEnd:
            logger.debug(f"{code} set to {matchEnd.group(1)}")
            result = matchEnd.group(1)
        return result
    else:
        logger.error(f"{code} contains '…' and code not in {mentionInFile[code]}")
        return result

def incrementer_section(match):
    """Incrémenter le dernier chiffre d'une section"""
    numero = int(match.group(3))
    nouveau = numero + 1
    return f"{match.group(1)}{match.group(2)}{nouveau}"


def main():
    """Fonction principale"""


    pictogrammes_identifies=[]

    if len(sys.argv) < 2:
        logger.error("Usage: python extract_pdf.py <pathPDF> [<fichier_etiq.xls>] [key_tmp]")
        sys.exit(1)
    pathPDF = sys.argv[1]
    logger.debug(f"Traitement du fichier PDF....: {pathPDF}")
    global dictMention
    logger.debug(f"Chargement du distionnaire depuis : {config['PATHS']['pathMention']}")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    dictMention = excel_to_dict(os.path.join(script_dir,config['PATHS']['pathMention']))


    fileSticker = sys.argv[2] if len(sys.argv) > 2 else ""
    key_tmp = sys.argv[3] if len(sys.argv) > 3 else config['SETTINGS']['suffixe_tmp']
    # Vérifier que le fichier existe
    if not Path(pathPDF).exists():
        logger.debug(f"❌ Le fichier '{pathPDF}' n'existe pas")
        sys.exit(1)

    logger.debug(f"🚀 Début de l'extraction de: {pathPDF}\n")

    # Extraire le texte
    results = extraire_texte_pdf(pathPDF)
    if results:
        logger.info(f"\n📝 Extracted: {results}")
        if results['pictos'] == []:
            logger.info("No pictos named ... Try to find out from pictures ")
            if (config['SETTINGS']['RECOGNIZE_PICTURES']):
                pictogrammes_identifies = analyser_fds(pathPDF)
            for picto in pictogrammes_identifies:
                logger.info(f"picto from imùage {picto}")
                if picto['code'] not in results['pictos']:
                    results['pictos'].append(picto['code'])

        logger.info(f"\n📝 Final: {results}")
        sheetName=''
        if (x := re.search(r"(.*)\.pdf$", os.path.basename(pathPDF), re.IGNORECASE)):
            sheetName=x.group(1)
        #SI nom de fichie etiquette positionnée on va mettre à jour toutes les lignes qui font référence à ce FDS
        # Sinon on va créer un onglet dans pathFdsExcel="Datas/FdsExcelNoAPI.xlsx"
        if fileSticker and not fileSticker == "":
            logger.debug(f"Report fds in file etiquette {fileSticker}")
            update_sticker_file(results, sheetName, fileSticker,key_tmp)
        else:
            logger.debug(f"Report fds in file sheet {sheetName} of file {config['PATHS']['pathFdsExcel']}")
            write_fds(results,sheetName)
    else:
        logger.error("❌ Aucun texte n'a pu être extrait")
        sys.exit(1)

    logger.info("\n✨ Extraction terminée!")


if __name__ == "__main__":
    main()